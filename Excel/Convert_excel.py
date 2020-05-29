# python 3.7.1
# Oloh 2018
# Convert xls to xlsx

"""Yet to create"""

import os
import glob
from pathlib import Path
import xlrd
from openpyxl.workbook import Workbook

XLS = '.xls'
XLSX = '.xlsx'

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    if not dst_file_path.exists():
        book_xlsx.save(dst_file_path)

print(os.getcwd())
# later on require argument
# os.chdir('s:/MB_M_F_A2_SW/SW_MAIN_CO/SW/Safety/CallTreeAnalysis/MFA')
excel_files = glob.glob('*'+XLS)
# target_names = []

for elName in excel_files:
    # nameParts = elName.split('.')
    if elName.endswith(XLS):
        newName = elName.replace(XLS, XLSX)
        p_orig = Path(elName)
        p_new = Path(newName)
        cvt_xls_to_xlsx(p_orig, p_new)


# for elName in excel_files:
#     name_part = elName[:len(elName)-4]
#     newName = name_part + '.' + elName[len(elName)-4:]
#     p_orig = Path(elName)
#     p_new = Path(newName)
#     p_orig.rename(p_new)
